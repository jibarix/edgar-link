"""Build a multi-peer comparables workbook from data/company_index.json.

Universe is selected by filtering the local company classification index
(SIC + optional name / subindustry / country filters); no anchor company
is required. For every peer in the resulting set the script pulls SEC
Company Facts, runs the derived-metric registry, and writes a single
styled Excel workbook containing:

    Universe         -- one row per peer with the classification fields
                        from data/company_index.json (name, ticker, CIK,
                        SIC, industry, subindustry, country/state of
                        incorporation, dominant revenue country, geo
                        breakdown summary, latest annual period).
    Metrics          -- peers x (metric x relative fiscal period) matrix.
                        Columns are FY 0 / FY -1 / FY -2 ... so peers
                        with different fiscal calendars line up. The
                        actual period dates per peer live on Universe.
    Screening_24col  -- CapIQ Company Screening Report mirror, single
                        point-in-time snapshot per peer (LTM where
                        possible, else FY-aligned annual).
    Screening_36col  -- 24-col layout + status + 4 forward-estimate cols
                        (blank by design; no EDGAR equivalent) + 6
                        trailing LTM revenue columns.
    <PEER>           -- one sheet per peer with BS / IS / CF normalized
                        line items stacked vertically, rows = display
                        name, columns = the peer's own period dates.
    About            -- methodology, filters used, blank-by-design notes.

Universe selection:

    --sic CODE [CODE ...]                one or more SIC codes (required)
    --name-substring TEXT [TEXT ...]     keep only peers whose name
                                         contains TEXT (case-insensitive,
                                         OR across multiple values)
    --exclude-subindustry TEXT [...]     drop peers whose subindustry
                                         contains TEXT (case-insensitive)
    --exclude-name TEXT [...]            drop peers whose name contains
                                         TEXT (case-insensitive)
    --country-inc XX                     ISO2 filter on country_inc
    --revenue-country XX                 ISO2 filter on dominant revenue
                                         country (only set when one
                                         country crosses 50% of segmented
                                         revenue)
    --limit N                            cap the peer count after sorting

Output / behaviour:

    --metrics SLUG[,SLUG...]             override default metric list
    --period-type annual|quarterly       default annual
    --num-periods N                      default 5 (annual) / 8 (quarterly)
    --as-of YYYY-MM-DD                   point-in-time anchor for the
                                         CapIQ snapshot sheets. Default:
                                         today.
    --no-capiq-layout                    skip the Screening_24col /
                                         Screening_36col sheets (faster:
                                         no per-peer quarterly parse,
                                         no submissions metadata).
    --no-beta                            skip 5Y monthly beta vs S&P
                                         500 for the CapIQ sheets.
                                         Default is on (fail-soft: a
                                         Yahoo Finance failure blanks
                                         the columns rather than
                                         aborting the build).
    --extensions                         merge captive-finance extension
                                         XBRL for OEM lenders (Deere /
                                         CNH etc.). Adds extra HTTP
                                         calls per peer.
    --output PATH                        default
                                         output/comps_<label>_<period>_<YYYYMMDD>.xlsx
    --dry-run                            print the peer universe and
                                         exit; no SEC calls.

Live runs require EDGAR_IDENTITY in the environment. --dry-run does not.
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

# Make the repo importable when invoked as a bare script.
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.settings import BASE_DIR, DEFAULT_OUTPUT_DIR  # noqa: E402
from edgar import metrics as edgar_metrics  # noqa: E402
from edgar._extension_mappings import EQUIPMENT_FINANCE_RULES  # noqa: E402
from edgar.company_lookup import format_cik, get_company_tickers  # noqa: E402
from edgar.filing_retrieval import FilingRetrieval  # noqa: E402
from edgar.metrics.ltm import build_ltm_statement  # noqa: E402
from edgar.xbrl_parser import XBRLParser  # noqa: E402

logger = logging.getLogger("build_comps")

INDEX_PATH = Path(BASE_DIR) / "data" / "company_index.json"

# Lookback buffer on top of the user's requested num_periods so trailing
# metrics like *_cagr_3y / *_growth still have prior periods to walk.
LOOKBACK_BUFFER = 4

# Default metric set. Keep small; users override via --metrics. Chosen
# to be informative for most SIC buckets, with inventory / working-cap
# metrics that matter for goods retailers (dealers, distributors).
DEFAULT_METRICS = [
    # scale + growth
    "revenue", "revenue_growth", "revenue_cagr_3y",
    # margins
    "gross_profit", "gross_profit_margin",
    "ebit", "ebit_margin",
    "ebitda", "ebitda_margin",
    "ni", "ni_margin",
    "fcf", "fcf_margin",
    # working capital
    "inventory_turnover", "days_inventory_out",
    "days_sales_out", "days_payables_out", "cash_conversion_cycle",
    # leverage / liquidity
    "current_ratio", "quick_ratio",
    "debt_to_equity", "debt_to_capital", "financial_leverage",
    # returns
    "roa", "roe", "roic", "asset_turnover",
]

# Statement-section order for per-peer drilldown sheets. Categories
# match the closed 9-pair vocabulary the parser stamps on each metric
# (BS = Assets/Liabilities/Equity, IS = Revenue/Income/EPS,
# CF = Operating/Investing/Financing).
PEER_SECTIONS: list[tuple[str, tuple[str, ...]]] = [
    ("Balance Sheet",       ("Assets", "Liabilities", "Equity")),
    ("Income Statement",    ("Revenue", "Income", "EPS")),
    ("Cash Flow Statement", ("OperatingCashFlow", "InvestingCashFlow",
                             "FinancingCashFlow")),
]

# Excel sheet-name rules: <=31 chars, no \ / ? * [ ] :
_SHEET_BAD = re.compile(r"[\\/?*\[\]:]")

# ── styling constants ─────────────────────────────────────────────────
# Header band mirrors the CapIQ Company Screening Report look-and-feel:
# bold white text on dark navy, wrap + top-align so long column labels
# render readably without manual row-height tuning. Pattern matches the
# archived generate_comps_workbook.py for visual consistency with prior
# deliverables.
_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_ALIGN = Alignment(wrap_text=True, vertical="top", horizontal="left")
_SECTION_FONT = Font(bold=True, italic=True)

# Number formats. Accounting-style currency keeps negatives in parens
# and zero as "-"; ratios render as percent; turnover ratios as Nx.
_FMT_USD = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
_FMT_USD_MM = '_-* #,##0.0_-;-* #,##0.0_-;_-* "-"??_-;_-@_-'
_FMT_RATIO_PCT = '0.0%;(0.0%);"-"'
_FMT_TURNOVER = '#,##0.00"x"'
_FMT_DAYS = '#,##0'
_FMT_FLOAT = '#,##0.000'


# ── The two native CapIQ Screening structures (column STRUCTURE only,
#    captured once from the vendor files; nothing is read from them at
#    runtime). 24-col is the prefix of 36-col. ─────────────────────────
_HDR_24 = [
    "Company Name", "Exchange:Ticker", "Company Type",
    "Exchanges [Primary Listing]", "Exchanges [Secondary Listings]",
    "Industry Classifications", "Geographic Locations",
    "Country/Region of Incorporation", "Primary Industry", "Primary Sector",
    "Short Business Description", "5 Year Beta [Latest]",
    "5 Year Beta R-Squared [Latest]",
    "Total Debt [Latest Annual] ($USDmm)",
    "Cash And Equivalents [Latest Annual] ($USDmm)",
    "Total Revenue [LTM] ($USDmm)",
    "Interest Expense [LTM] ($USDmm)",
    "Gross Profit [LTM] ($USDmm)",
    "EBIT [LTM] ($USDmm)",
    "EBITDA [LTM] ($USDmm)",
    "Net Income [LTM] ($USDmm)",
    "Change in Net Working Capital [LTM] ($USDmm)",
    "Capital Expenditure [LTM] ($USDmm)",
    "Depreciation & Amort. [LTM] ($USDmm)",
]
_HDR_36 = _HDR_24 + [
    "Company Status",
    "Est. Annual Revenue Growth - 1 Yr % [Latest] (%)",
    "Est. Annual Revenue Growth - 2 Yr % [Latest] (%)",
    "Est. Annual Revenue Growth - 1 Yr % [Latest]",
    "Est. Annual Revenue Growth - 2 Yr % [Latest]",
    "Total Revenue [LTM - 1] ($USDmm)",
    "Total Revenue [LTM - 2] ($USDmm)",
    "Total Revenue [LTM - 3] ($USDmm)",
    "Total Revenue [LTM - 4] ($USDmm)",
    "Total Revenue [LTM - 5] ($USDmm)",
    "Total Revenue [LTM - 6] ($USDmm)",
]
# CapIQ forward analyst estimates have no EDGAR equivalent. The four
# "Est. Annual Revenue Growth" columns are intentionally left blank;
# the About sheet documents this. Indices into _HDR_36 (0-based):
_BLANK_BY_DESIGN = {25, 26, 27, 28}

# First numeric column in the CapIQ layout (1-based, openpyxl convention).
# Cols 1..11 are text; col 12 onward is numeric data.
_CAPIQ_FIRST_NUM_COL = 12


# ── index loading + filtering ─────────────────────────────────────────


def _load_index() -> dict:
    if not INDEX_PATH.exists():
        raise SystemExit(
            f"company_index.json not found at {INDEX_PATH}. Build it first:\n"
            f"    python -m edgar.company_classifier --build"
        )
    with INDEX_PATH.open(encoding="utf-8") as fh:
        return json.load(fh)


def _apply_filters(index: dict, args) -> list[tuple[str, dict]]:
    """Return [(cik, entry), ...] for peers passing every filter, sorted
    by name. CIK key in the returned tuple is the raw (non-padded) form
    used inside the index; zero-pad with format_cik() before SEC calls.
    """
    sics = set(str(s).zfill(4) for s in args.sic)
    name_subs = [s.lower() for s in (args.name_substring or [])]
    excl_sub = [s.lower() for s in (args.exclude_subindustry or [])]
    excl_name = [s.lower() for s in (args.exclude_name or [])]

    out: list[tuple[str, dict]] = []
    for cik, entry in index.items():
        if str(entry.get("sic") or "").zfill(4) not in sics:
            continue
        name = (entry.get("name") or "").lower()
        sub = (entry.get("subindustry") or "").lower()
        if name_subs and not any(s in name for s in name_subs):
            continue
        if any(s in sub for s in excl_sub):
            continue
        if any(s in name for s in excl_name):
            continue
        if args.country_inc and (entry.get("country_inc") or "").upper() != args.country_inc.upper():
            continue
        if args.revenue_country and (entry.get("revenue_country") or "").upper() != args.revenue_country.upper():
            continue
        out.append((cik, entry))

    out.sort(key=lambda kv: (kv[1].get("name") or "").upper())
    if args.limit and len(out) > args.limit:
        out = out[: args.limit]
    return out


def _build_cik_to_ticker() -> dict[str, str]:
    """Reverse map: zero-padded CIK -> ticker. Best-effort; missing peers
    just get an empty ticker.
    """
    data = get_company_tickers()
    out: dict[str, str] = {}
    for entry in data["by_ticker"].values():
        out.setdefault(entry["cik"], entry["ticker"])
    return out


# ── peer pull + metric compute ────────────────────────────────────────


def _pull_peer_facts(
    cik_padded: str,
    filings: FilingRetrieval,
    parser: XBRLParser,
    use_extensions: bool,
) -> dict | None:
    """Pull raw Company Facts (with optional extension merge) for one
    peer. Returns the augmented facts blob or None on failure.
    """
    facts = filings.get_company_facts(cik_padded)
    if not facts:
        return None
    if use_extensions:
        try:
            md = filings.get_filing_metadata(cik_padded, "10-K", limit=6)
            if md:
                parser.augment_with_extensions(
                    facts, filings, cik_padded, md, EQUIPMENT_FINANCE_RULES,
                )
        except Exception as e:  # noqa: BLE001
            logger.warning("extension merge failed for %s: %s", cik_padded, e)
    return facts


def _parse_normalized(
    facts: dict,
    parser: XBRLParser,
    period_type: str,
    num_periods: int,
) -> dict | None:
    """Run XBRLParser.parse_company_facts and validate output."""
    normalized = parser.parse_company_facts(
        facts,
        statement_type="ALL",
        period_type=period_type,
        num_periods=num_periods,
    )
    if not normalized or not normalized.get("periods"):
        return None
    return normalized


def _compute_metric_series(
    normalized: dict, slugs: list[str]
) -> dict[str, dict[str, float | None]]:
    """slug -> {period: value} for one peer."""
    ns = edgar_metrics.NormalizedStatement(normalized)
    out: dict[str, dict[str, float | None]] = {}
    for slug in slugs:
        spec = edgar_metrics.REGISTRY.get(slug)
        if spec is None:
            out[slug] = {p: None for p in ns.periods}
            continue
        try:
            out[slug] = spec.fn(ns)
        except Exception as e:  # noqa: BLE001  -- per-peer safety
            logger.warning("metric %s failed for one peer: %s", slug, e)
            out[slug] = {p: None for p in ns.periods}
    return out


# ── point-in-time CapIQ snapshot ──────────────────────────────────────
# Mirrors reconcile_capiq_screening.compute_edgar's resolution logic:
# - For Dec-year-end filers, the most recent annual <= as_of is fine.
# - For non-Dec filers, prefer an LTM rollup ending closest to as_of so
#   the snapshot is fiscal-calendar-aligned to the screen date rather
#   than the issuer's odd FY end.


def _pick_annual_period(periods: list[str], as_of: str) -> str | None:
    """Annual period whose fiscal year matches `as_of`, else the most
    recent annual period <= as_of, else the most recent annual.
    """
    if not periods:
        return None
    yr = as_of[:4]
    for p in periods:
        if p[:4] == yr:
            return p
    for p in periods:
        if p <= as_of:
            return p
    return periods[0]


def _capiq_snapshot(
    annual_norm: dict,
    qtr_norm: dict | None,
    as_of: str,
) -> tuple[edgar_metrics.NormalizedStatement, str, str] | None:
    """Return (statement, latest_period, ptype) where ptype is "ltm" or
    "annual". For non-Dec filers with quarterly history available, build
    an LTM statement ending at the most recent quarterly period <= as_of.
    Otherwise fall back to the FY-aligned annual snapshot.
    """
    a_periods = annual_norm.get("periods", [])
    if not a_periods:
        return None
    # Dec-year-end OR latest annual == as_of: annual is exact.
    if a_periods[0] == as_of or a_periods[0][5:] == "12-31":
        stmt = edgar_metrics.NormalizedStatement(annual_norm)
        latest = _pick_annual_period(stmt.periods, as_of)
        return (stmt, latest, "annual") if latest else None
    # Try LTM if we have a quarterly slice.
    if qtr_norm:
        ltm = build_ltm_statement(annual_norm, qtr_norm, as_of, None)
        if ltm is not None:
            stmt, latest = ltm
            return stmt, latest, "ltm"
    # Fall back to annual.
    stmt = edgar_metrics.NormalizedStatement(annual_norm)
    latest = _pick_annual_period(stmt.periods, as_of)
    return (stmt, latest, "annual") if latest else None


def _compute_capiq_row(
    stmt: edgar_metrics.NormalizedStatement,
    latest: str,
) -> dict[str, float | None]:
    """Resolve the CapIQ-layout numeric block (raw USD) for one peer."""

    def m(slug: str) -> float | None:
        spec = edgar_metrics.REGISTRY.get(slug)
        return spec.fn(stmt).get(latest) if spec else None

    rev = m("revenue")

    def from_margin(slug: str) -> float | None:
        """CapIQ "Interest Expense [LTM]" / "D&A [LTM]" need raw dollars;
        the registry exposes them as margins (× revenue → dollars)."""
        mg = m(slug)
        return mg * rev if mg is not None and rev is not None else None

    # ΔNWC = NWC_t − NWC_{t−1}. NWC is a registered metric slug (in
    # REGISTRY), not a CONCEPT_CHAINS entry, so route through the
    # registry rather than stmt.get(...) which would KeyError.
    nwc_spec = edgar_metrics.REGISTRY.get("nwc")
    nwc_series = nwc_spec.fn(stmt) if nwc_spec else {}
    prior = stmt.prior_period(latest)
    dnwc = None
    if prior is not None:
        cur = nwc_series.get(latest)
        prv = nwc_series.get(prior)
        if cur is not None and prv is not None:
            dnwc = cur - prv

    # Capex = raw concept payment. `capex` is a logical concept in
    # CONCEPT_CHAINS (resolves to PaymentsToAcquirePPE et al.). Cash-flow
    # outflows are stored positive in XBRL; report as positive to match
    # CapIQ convention.
    try:
        capex_series = stmt.get("capex")
    except KeyError:
        capex_series = {}
    capex = capex_series.get(latest)
    if capex is not None:
        capex = abs(capex)

    return {
        "revenue":          rev,
        "gross_profit":     m("gross_profit"),
        "ebit":             m("ebit"),
        "ebitda":           m("ebitda"),
        "ni":               m("ni"),
        "interest_expense": from_margin("interest_exp_margin"),
        "da":               from_margin("da_margin"),
        "cash":             m("cash_and_st_investments"),
        "total_debt":       m("total_debt"),
        "dnwc":             dnwc,
        "capex":            capex,
    }


def _trailing_annual_revenue(
    annual_norm: dict,
    as_of: str,
    n: int = 7,
) -> list[float | None]:
    """Most-recent N annual revenues <= as_of (raw USD), newest-first."""
    ns = edgar_metrics.NormalizedStatement(annual_norm)
    eligible = [p for p in ns.periods if p <= as_of]
    if not eligible:
        return []
    spec = edgar_metrics.REGISTRY.get("revenue")
    if spec is None:
        return []
    series = spec.fn(ns)
    return [series.get(p) for p in eligible[:n]]


def _submissions_meta(cik_padded: str, filings: FilingRetrieval) -> dict:
    """Best-effort SEC submissions metadata for the CapIQ text block.
    Missing fields → empty string (never faked).
    """
    out = {
        "name": "", "exch": "", "exch_all": [], "sic_desc": "",
        "geo": "", "inc": "", "status": "Operating",
    }
    try:
        sub = filings.get_company_submissions(cik_padded)
        if not sub:
            return out
        out["name"] = sub.get("name") or ""
        out["sic_desc"] = sub.get("sicDescription", "") or ""
        out["inc"] = sub.get("stateOfIncorporation", "") or ""
        ex = sub.get("exchanges") or []
        out["exch_all"] = [str(x) for x in ex]
        out["exch"] = out["exch_all"][0] if out["exch_all"] else ""
        biz = (sub.get("addresses") or {}).get("business") or {}
        city = (biz.get("city") or "").strip()
        region = (biz.get("stateOrCountry") or "").strip()
        out["geo"] = ", ".join(p for p in (city, region) if p)
        if str(sub.get("entityType", "")).lower() in ("", "operating"):
            out["status"] = "Operating"
    except Exception as e:  # noqa: BLE001
        logger.warning("submissions lookup failed for %s: %s", cik_padded, e)
    return out


# ── workbook writers (openpyxl) ───────────────────────────────────────


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _HDR_ALIGN


def _safe_sheet_name(label: str, used: set[str]) -> str:
    clean = _SHEET_BAD.sub("_", label)[:31] or "peer"
    if clean not in used:
        used.add(clean)
        return clean
    for i in range(2, 100):
        suffix = f"_{i}"
        cand = (clean[: 31 - len(suffix)] + suffix)
        if cand not in used:
            used.add(cand)
            return cand
    used.add(clean)
    return clean


def _format_geo(geo: dict | None) -> str:
    if not geo:
        return ""
    items = sorted(geo.items(), key=lambda kv: -kv[1])
    return "; ".join(f"{c}: {p:.1f}%" for c, p in items)


def _unit_for(slug: str) -> str:
    spec = edgar_metrics.REGISTRY.get(slug)
    return (spec.unit or "") if spec else ""


def _number_format_for(unit: str) -> str:
    if unit == "USD":
        return _FMT_USD
    if unit == "ratio":
        return _FMT_RATIO_PCT
    if unit == "x":
        return _FMT_TURNOVER
    if unit == "days":
        return _FMT_DAYS
    return _FMT_FLOAT


def _write_universe(
    wb: Workbook,
    peers: list[tuple[str, dict]],
    cik_to_ticker: dict[str, str],
    peer_pulls: dict[str, dict],
    num_periods: int,
) -> None:
    ws = wb.create_sheet("Universe")
    headers = [
        "CIK", "Ticker", "Name", "SIC", "Industry", "Subindustry",
        "Country_Inc", "State_Inc", "Revenue_Country", "Revenue_Pct",
        "Geo_Breakdown", "Index_Period", "Periods_Loaded", "Pull_Status",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for cik_raw, entry in peers:
        cik_padded = format_cik(cik_raw)
        pull = peer_pulls.get(cik_padded) or {}
        peer_periods = (pull.get("periods") or [])[:num_periods]
        ws.append([
            cik_padded,
            cik_to_ticker.get(cik_padded, ""),
            entry.get("name") or "",
            str(entry.get("sic") or ""),
            entry.get("industry") or "",
            entry.get("subindustry") or "",
            entry.get("country_inc") or "",
            entry.get("state_inc") or "",
            entry.get("revenue_country") or "",
            entry.get("revenue_pct"),
            _format_geo(entry.get("geo_breakdown")),
            entry.get("period") or "",
            ", ".join(peer_periods) if peer_periods else "(no data)",
            "ok" if pull.get("periods") else "skipped (no facts)",
        ])

    # Revenue_Pct (col 10) as percent.
    pct_col = headers.index("Revenue_Pct") + 1
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=pct_col)
        if isinstance(cell.value, (int, float)):
            # Index stores 0-100; render as percent of 100.
            cell.value = cell.value / 100.0
            cell.number_format = _FMT_RATIO_PCT

    # Column widths: name + breakdown wide, ratios narrow.
    widths = {
        1: 12, 2: 10, 3: 36, 4: 8, 5: 22, 6: 28, 7: 10, 8: 10,
        9: 12, 10: 12, 11: 36, 12: 12, 13: 32, 14: 22,
    }
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_metrics_matrix(
    wb: Workbook,
    peers: list[tuple[str, dict]],
    cik_to_ticker: dict[str, str],
    peer_pulls: dict[str, dict],
    metric_slugs: list[str],
    num_periods: int,
) -> None:
    """Rows: peer; Columns: peer label + (metric, relative fiscal index).

    Periods are aligned by relative fiscal position so peers with
    different fiscal year-ends line up: FY 0 = each peer's most recent
    period, FY -1 = the one before, etc.
    """
    ws = wb.create_sheet("Metrics")
    rel_labels = [f"FY {0 if i == 0 else -i}" for i in range(num_periods)]

    # Two header rows: row 1 = metric (merged across rel periods),
    # row 2 = rel-period label. Col A = peer label.
    row1 = ["Peer"]
    row2 = [""]
    for slug in metric_slugs:
        unit = _unit_for(slug)
        header = slug + (f"  [{unit}]" if unit else "")
        row1.append(header)
        row1.extend([""] * (num_periods - 1))
        row2.extend(rel_labels)
    ws.append(row1)
    ws.append(row2)

    # Merge metric label across its N rel-period columns.
    for i, _slug in enumerate(metric_slugs):
        start_col = 2 + i * num_periods
        end_col = start_col + num_periods - 1
        if num_periods > 1:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col,
            )
    _style_header(ws, 1, 1 + len(metric_slugs) * num_periods)
    _style_header(ws, 2, 1 + len(metric_slugs) * num_periods)
    ws.cell(row=1, column=1).alignment = Alignment(
        horizontal="left", vertical="top",
    )

    # Body.
    for cik_raw, entry in peers:
        cik_padded = format_cik(cik_raw)
        ticker = cik_to_ticker.get(cik_padded, "")
        label = f"{ticker} ({cik_padded})" if ticker else cik_padded
        pull = peer_pulls.get(cik_padded)
        row: list = [label]
        if not pull or not pull.get("periods"):
            row.extend([None] * len(metric_slugs) * num_periods)
            ws.append(row)
            continue
        series_by_slug = _compute_metric_series(pull, metric_slugs)
        peer_periods = pull["periods"][:num_periods]
        for slug in metric_slugs:
            series = series_by_slug.get(slug, {})
            for i in range(num_periods):
                p = peer_periods[i] if i < len(peer_periods) else None
                v = series.get(p) if p else None
                row.append(v)
        ws.append(row)

    # Per-metric number format on the body rows.
    for i, slug in enumerate(metric_slugs):
        fmt = _number_format_for(_unit_for(slug))
        start_col = 2 + i * num_periods
        end_col = start_col + num_periods - 1
        for r in range(3, ws.max_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).number_format = fmt

    # Column widths.
    ws.column_dimensions["A"].width = 28
    for c in range(2, 1 + len(metric_slugs) * num_periods + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = (
        f"A2:{get_column_letter(1 + len(metric_slugs) * num_periods)}2"
    )


def _write_screening(
    wb: Workbook,
    sheet_name: str,
    hdr: list[str],
    rows: list[tuple[str, list]],
) -> None:
    """Generic CapIQ-layout writer for both Screening_24col and 36col."""
    ws = wb.create_sheet(sheet_name)
    ncols = len(hdr)
    ws.append(hdr)
    _style_header(ws, 1, ncols)

    # Body.
    for _tkr, full in rows:
        ws.append(full[:ncols])

    # Number formats: USD-millions from col 14 onward (col 12 = beta,
    # col 13 = R²: ratios). Forward-estimate cols 26-29 are blank by
    # design; they get the same numeric fmt for any future fill-ins.
    for r in range(2, ws.max_row + 1):
        # Beta + R²: 3-decimal float.
        for c in (12, 13):
            if c <= ncols:
                ws.cell(row=r, column=c).number_format = _FMT_FLOAT
        # Currency block ($USDmm).
        for c in range(14, ncols + 1):
            # Skip the forward-estimate text/blank cols if 36-wide.
            if ncols >= 30 and (c - 1) in _BLANK_BY_DESIGN:
                continue
            # Col 26 (Company Status) is text.
            if ncols >= 25 and c == 25:
                continue
            ws.cell(row=r, column=c).number_format = _FMT_USD_MM

    # Column widths: name-ish wide, others compact.
    for ci in range(1, ncols + 1):
        if ci in (1, 6, 9, 11):
            w = 40
        elif ci in (2, 4, 7, 8, 10):
            w = 26
        else:
            w = 18
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _build_capiq_row(
    cik_padded: str,
    ticker: str,
    meta: dict,
    classification: dict,
    snap: tuple | None,
    beta: tuple[float | None, float | None] | None,
    trailing_rev_mm: list[float | None],
) -> list:
    """Assemble one 36-wide row; 24-col sheet slices [:24]."""
    row: list = [None] * len(_HDR_36)
    row[0] = meta.get("name") or classification.get("name") or ""
    row[1] = (f"{meta.get('exch','')}:{ticker}"
              if meta.get("exch") and ticker else ticker or cik_padded)
    row[2] = "Public Company"
    row[3] = meta.get("exch") or ""
    row[4] = ", ".join(meta.get("exch_all") or [])
    row[5] = meta.get("sic_desc") or classification.get("subindustry") or ""
    row[6] = meta.get("geo") or ""
    row[7] = meta.get("inc") or classification.get("state_inc") or ""
    row[8] = classification.get("subindustry") or meta.get("sic_desc") or ""
    row[9] = classification.get("industry") or ""
    row[10] = ""  # SEC submissions carry no reliable prose description.
    if beta is not None:
        row[11], row[12] = beta
    if snap is not None:
        _stmt, _latest, _ptype, vals = snap
        row[13] = _mm(vals.get("total_debt"))
        row[14] = _mm(vals.get("cash"))
        row[15] = _mm(vals.get("revenue"))
        row[16] = _mm(vals.get("interest_expense"))
        row[17] = _mm(vals.get("gross_profit"))
        row[18] = _mm(vals.get("ebit"))
        row[19] = _mm(vals.get("ebitda"))
        row[20] = _mm(vals.get("ni"))
        row[21] = _mm(vals.get("dnwc"))
        row[22] = _mm(vals.get("capex"))
        row[23] = _mm(vals.get("da"))
    row[24] = meta.get("status") or "Operating"
    # 25-28: forward analyst estimates, blank by design.
    for j, v_mm in enumerate(trailing_rev_mm[1:7], start=29):
        if j < len(_HDR_36):
            row[j] = v_mm
    return row


def _mm(v: float | None) -> float | None:
    """Raw USD -> $mm. Pass None through."""
    return None if v is None else round(v / 1e6, 3)


def _write_peer_sheet(
    wb: Workbook,
    sheet_name: str,
    pull: dict,
    num_periods: int,
) -> None:
    """One sheet per peer: BS / IS / CF stacked vertically, columns are
    the peer's own period dates.
    """
    ws = wb.create_sheet(sheet_name)
    periods = pull["periods"][:num_periods]
    metrics = pull.get("metrics", {})

    by_cat: dict[str, list[tuple[str, dict]]] = {}
    for key, meta in metrics.items():
        cat = meta.get("category") or ""
        by_cat.setdefault(cat, []).append((key, meta))

    headers = ["Line", *periods]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    body_start_row = 2
    for title, cats in PEER_SECTIONS:
        items: list[tuple[str, dict]] = []
        for c in cats:
            items.extend(by_cat.get(c, []))
        if not items:
            continue
        # Section header. Avoid a leading "=" -- Excel treats cells
        # starting with "=" as formulas and silently mangles them.
        ws.append([f"[ {title} ]"] + [None] * len(periods))
        section_row = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=section_row, column=c)
            cell.font = _SECTION_FONT
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
        # Stable order: parser's `order` puts subtotals where they
        # belong; ties break on display name.
        items.sort(key=lambda kv: (
            (kv[1].get("order") or 50),
            (kv[1].get("display_name") or kv[0]).lower(),
        ))
        for key, meta in items:
            display = meta.get("display_name") or meta.get("tag") or key
            values = meta.get("values", {})
            ws.append([display] + [values.get(p) for p in periods])

    if ws.max_row == 1:
        ws.append(["(no statement data parsed)"] + [None] * len(periods))

    # Number format USD on all numeric cells.
    for r in range(body_start_row, ws.max_row + 1):
        for c in range(2, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = _FMT_USD

    ws.column_dimensions["A"].width = 44
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    ws.freeze_panes = "B2"


def _write_about(
    wb: Workbook,
    args,
    peer_count: int,
    pulled_count: int,
    metric_slugs: list[str],
    as_of: str,
    capiq_on: bool,
    extensions_on: bool,
    beta_on: bool,
) -> None:
    ws = wb.create_sheet("About")
    notes: list[tuple[str, str]] = [
        ("EDGAR comparables workbook — methodology", ""),
        ("", ""),
        ("Universe selection", ""),
        ("  SIC code(s)", ", ".join(str(s) for s in args.sic)),
        ("  Name substring(s)",
         ", ".join(args.name_substring or []) or "(none)"),
        ("  Exclude subindustry",
         ", ".join(args.exclude_subindustry or []) or "(none)"),
        ("  Exclude name",
         ", ".join(args.exclude_name or []) or "(none)"),
        ("  Country of incorporation", args.country_inc or "(any)"),
        ("  Revenue country", args.revenue_country or "(any)"),
        ("  Peer cap", str(args.limit) if args.limit else "(none)"),
        ("  Peers resolved", f"{pulled_count}/{peer_count} returned EDGAR data"),
        ("", ""),
        ("Metrics matrix", ""),
        ("  Metric slugs", ", ".join(metric_slugs)),
        ("  Period type", args.period_type),
        ("  Periods", str(args.num_periods)),
        ("  Period alignment",
         "FY-relative: FY 0 = each peer's most recent period; FY -k = "
         "the k-th prior period in that peer's own fiscal calendar. "
         "Periods do NOT line up across peers — Universe sheet lists "
         "each peer's actual period dates."),
        ("", ""),
        ("Screening sheets (CapIQ layout)",
         "included" if capiq_on else "skipped (--no-capiq-layout)"),
        ("  As-of (point-in-time)", as_of),
        ("  Period basis",
         "LTM = trailing 4 quarters with fiscal period-end <= as-of; "
         "else most recent FY with period-end <= as-of. Mirrors the "
         "resolution used by the archived reconcile_capiq_screening "
         "tool against vendor CapIQ files."),
        ("  Units", "$USDmm (engine returns raw USD; /1e6 here)."),
        ("  Captive-finance extensions",
         "on (EQUIPMENT_FINANCE_RULES merged)" if extensions_on
         else "off (use --extensions to merge OEM-lender extension "
              "XBRL — Deere post-FY22 etc.)"),
        ("  5Y monthly beta",
         "on (vs ^GSPC log returns)" if beta_on
         else "off (use --include-beta to compute)"),
        ("", ""),
        ("Blank by design (no EDGAR equivalent — never fabricated)", ""),
        ("  Cols 26-29 (Screening_36col)",
         "CapIQ forward analyst revenue-growth estimates."),
        ("  Short Business Description",
         "SEC submissions carry no reliable prose description."),
        ("  Market Capitalization",
         "Requires as-of price x shares; not pulled in this build "
         "(EDGAR has no price feed). Left blank rather than approximated."),
        ("", ""),
        ("Data source",
         "100% SEC EDGAR XBRL via the edgar-connect engine."),
        ("Generated", date.today().isoformat()),
    ]
    for k, v in notes:
        ws.append([k, v])
    ws["A1"].font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 96
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1)
        b = ws.cell(row=r, column=2)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        if a.value and not a.value.startswith(" ") and not b.value:
            a.font = Font(bold=True)


# ── CLI ────────────────────────────────────────────────────────────────


def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="build_comps",
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--sic", nargs="+", required=True,
                   help="One or more SIC codes (e.g. 5500 5511).")
    p.add_argument("--name-substring", nargs="+",
                   help="Keep only peers whose name contains any of these "
                        "substrings (case-insensitive, OR).")
    p.add_argument("--exclude-subindustry", nargs="+",
                   help="Drop peers whose subindustry contains any of "
                        "these substrings.")
    p.add_argument("--exclude-name", nargs="+",
                   help="Drop peers whose name contains any of these "
                        "substrings.")
    p.add_argument("--country-inc",
                   help="ISO2 country-of-incorporation filter.")
    p.add_argument("--revenue-country",
                   help="ISO2 dominant-revenue-country filter (only set "
                        "when one country crosses 50%% of segmented "
                        "revenue).")
    p.add_argument("--limit", type=int,
                   help="Cap the peer count.")
    p.add_argument("--metrics",
                   help="Comma-separated metric slugs. Defaults to a "
                        "curated registry subset.")
    p.add_argument("--period-type", choices=("annual", "quarterly"),
                   default="annual",
                   help="Default: annual.")
    p.add_argument("--num-periods", type=int,
                   help="Periods to display. Default 5 (annual) / 8 "
                        "(quarterly).")
    p.add_argument("--as-of", dest="as_of",
                   help="YYYY-MM-DD point-in-time anchor for the CapIQ "
                        "snapshot sheets. Default: today.")
    p.add_argument("--no-capiq-layout", dest="capiq_layout",
                   action="store_false", default=True,
                   help="Skip the CapIQ Screening_24col / Screening_36col "
                        "sheets (faster).")
    p.add_argument("--no-beta", dest="include_beta",
                   action="store_false", default=True,
                   help="Skip the 5Y monthly beta vs ^GSPC computation "
                        "for CapIQ sheets. Default is on (fail-soft: "
                        "Yahoo Finance errors blank the columns rather "
                        "than aborting the build).")
    p.add_argument("--extensions", action="store_true",
                   help="Merge captive-finance extension XBRL "
                        "(EQUIPMENT_FINANCE_RULES). Extra HTTP per peer.")
    p.add_argument("--output",
                   help="Output workbook path. Default "
                        "output/comps_<label>_<period>_<YYYYMMDD>.xlsx.")
    p.add_argument("--dry-run", action="store_true",
                   help="Print the peer universe and exit; no SEC calls.")
    p.add_argument("--verbose", "-v", action="store_true",
                   help="Verbose logging.")
    return p


def _resolve_metrics(arg: str | None) -> list[str]:
    slugs = [s.strip() for s in arg.split(",")] if arg else list(DEFAULT_METRICS)
    slugs = [s for s in slugs if s]
    unknown = [s for s in slugs if s not in edgar_metrics.REGISTRY]
    if unknown:
        raise SystemExit(
            "Unknown metric slug(s): " + ", ".join(unknown)
            + "\n(use python -c \"from edgar.metrics import list_slugs; "
              "print(list_slugs())\" to enumerate)"
        )
    return slugs


def _default_output_path(args) -> Path:
    label = "sic" + "-".join(str(s).zfill(4) for s in args.sic)
    if args.name_substring:
        label += "_" + "-".join(re.sub(r"\W+", "", s).lower()
                                for s in args.name_substring)[:32]
    label = label[:64]
    stamp = datetime.now().strftime("%Y%m%d")
    return Path(DEFAULT_OUTPUT_DIR) / f"comps_{label}_{args.period_type}_{stamp}.xlsx"


def main(argv: list[str] | None = None) -> int:
    args = _build_arg_parser().parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )

    if args.num_periods is None:
        args.num_periods = 5 if args.period_type == "annual" else 8

    as_of = args.as_of or date.today().isoformat()
    try:
        date.fromisoformat(as_of)
    except ValueError:
        print(f"--as-of must be YYYY-MM-DD, got {as_of!r}", file=sys.stderr)
        return 2

    metric_slugs = _resolve_metrics(args.metrics)

    index = _load_index()
    peers = _apply_filters(index, args)

    if not peers:
        print("No peers matched the supplied filters.", file=sys.stderr)
        return 1

    cik_to_ticker = _build_cik_to_ticker() if not args.dry_run else {}

    # ── dry-run: print and exit ──
    if args.dry_run:
        print(f"Peer universe ({len(peers)} filers):")
        for cik_raw, entry in peers:
            cik_padded = format_cik(cik_raw)
            print(f"  {cik_padded}  SIC {entry.get('sic'):<6} "
                  f"{(entry.get('subindustry') or '')[:32]:32s}  "
                  f"{entry.get('name')}")
        print(f"\n(metrics: {len(metric_slugs)}, period_type: "
              f"{args.period_type}, num_periods: {args.num_periods}, "
              f"as_of: {as_of}, capiq_layout: {args.capiq_layout}, "
              f"extensions: {args.extensions}, beta: {args.include_beta})")
        return 0

    # ── live: enforce identity ──
    if not os.environ.get("EDGAR_IDENTITY"):
        print("EDGAR_IDENTITY is not set. Set it before running a live "
              "comps build:\n"
              "    $env:EDGAR_IDENTITY = \"Your Name your@email.com\"  "
              "(PowerShell)\n"
              "    export EDGAR_IDENTITY=\"Your Name your@email.com\"   "
              "(bash)",
              file=sys.stderr)
        return 2

    filings = FilingRetrieval()
    parser = XBRLParser()
    peer_pulls: dict[str, dict] = {}
    peer_facts: dict[str, dict] = {}
    peer_quarterly: dict[str, dict] = {}
    peer_meta: dict[str, dict] = {}
    peer_snapshots: dict[str, tuple | None] = {}
    peer_trailing_rev: dict[str, list[float | None]] = {}

    print(f"Pulling Company Facts for {len(peers)} peers"
          f"{' (with extensions)' if args.extensions else ''}...")
    for i, (cik_raw, entry) in enumerate(peers, 1):
        cik_padded = format_cik(cik_raw)
        name = entry.get("name") or cik_padded
        print(f"  [{i:>2}/{len(peers)}] {cik_padded}  {name}")
        try:
            facts = _pull_peer_facts(
                cik_padded, filings, parser, args.extensions,
            )
        except Exception as e:  # noqa: BLE001
            logger.warning("facts pull failed for %s: %s", cik_padded, e)
            facts = None
        if not facts:
            print(f"      ! no facts; peer will be empty in metrics matrix")
            peer_pulls[cik_padded] = {}
            continue
        peer_facts[cik_padded] = facts

        # Annual parse for the matrix + drilldown sheet.
        annual = _parse_normalized(
            facts, parser, args.period_type,
            args.num_periods + LOOKBACK_BUFFER,
        )
        peer_pulls[cik_padded] = annual or {}

        if args.capiq_layout:
            # Quarterly parse (only useful for LTM rollup of non-Dec
            # filers; cheap because facts are already in memory).
            try:
                qtr = parser.parse_company_facts(
                    facts, statement_type="ALL",
                    period_type="quarterly", num_periods=12,
                )
            except Exception as e:  # noqa: BLE001
                logger.warning("quarterly parse failed for %s: %s",
                               cik_padded, e)
                qtr = None
            peer_quarterly[cik_padded] = qtr or {}
            # Annual frame for the CapIQ snapshot uses a wider window so
            # _pick_annual_period can find the right FY <= as_of.
            annual_for_snap = annual
            if annual_for_snap and annual_for_snap.get("periods"):
                snap = _capiq_snapshot(annual_for_snap, qtr, as_of)
                if snap:
                    stmt, latest, ptype = snap
                    vals = _compute_capiq_row(stmt, latest)
                    peer_snapshots[cik_padded] = (stmt, latest, ptype, vals)
                    peer_trailing_rev[cik_padded] = (
                        _trailing_annual_revenue(annual_for_snap, as_of)
                    )
            peer_meta[cik_padded] = _submissions_meta(cik_padded, filings)

    # Optional beta batch (CapIQ layout only).
    beta_by_ticker: dict[str, tuple[float | None, float | None]] = {}
    if args.capiq_layout and args.include_beta:
        try:
            from edgar.metrics.beta import compute_peer_betas
            ts = [cik_to_ticker.get(format_cik(c), "")
                  for c, _ in peers]
            ts = [t for t in ts if t]
            print(f"Computing 5Y beta for {len(ts)} tickers (Yahoo)...")
            for pb in compute_peer_betas(ts, as_of=as_of):
                beta_by_ticker[pb.ticker.upper()] = (pb.beta, pb.r_squared)
        except Exception as e:  # noqa: BLE001
            print(f"  beta batch failed ({e.__class__.__name__}: {e}); "
                  f"beta columns will be blank")

    # Build CapIQ rows (anchor-less: sort by primary-period revenue desc).
    capiq_rows: list[tuple[str, list]] = []
    if args.capiq_layout:
        for cik_raw, entry in peers:
            cik_padded = format_cik(cik_raw)
            ticker = cik_to_ticker.get(cik_padded, "")
            meta = peer_meta.get(cik_padded, {})
            snap = peer_snapshots.get(cik_padded)
            beta = beta_by_ticker.get(ticker.upper()) if ticker else None
            trailing = peer_trailing_rev.get(cik_padded, [])
            trailing_mm = [_mm(v) for v in trailing]
            classification = {
                "name": entry.get("name") or "",
                "subindustry": entry.get("subindustry") or "",
                "industry": entry.get("industry") or "",
                "state_inc": entry.get("state_inc") or "",
            }
            row = _build_capiq_row(
                cik_padded, ticker, meta, classification,
                snap, beta, trailing_mm,
            )
            capiq_rows.append((ticker or cik_padded, row))
        # Sort by LTM revenue desc (col 16, index 15 = revenue), None last.
        capiq_rows.sort(key=lambda r: (
            r[1][15] is None, -(r[1][15] or 0.0),
        ))

    out_path = Path(args.output) if args.output else _default_output_path(args)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"\nWriting workbook: {out_path}")
    wb = Workbook()
    # Drop the default sheet that openpyxl creates.
    wb.remove(wb.active)

    _write_universe(
        wb, peers, cik_to_ticker, peer_pulls, args.num_periods,
    )
    _write_metrics_matrix(
        wb, peers, cik_to_ticker, peer_pulls, metric_slugs,
        args.num_periods,
    )
    if args.capiq_layout:
        _write_screening(wb, "Screening_24col", _HDR_24, capiq_rows)
        _write_screening(wb, "Screening_36col", _HDR_36, capiq_rows)
    used_sheet_names: set[str] = {
        "Universe", "Metrics", "Screening_24col", "Screening_36col",
    }
    for cik_raw, entry in peers:
        cik_padded = format_cik(cik_raw)
        pull = peer_pulls.get(cik_padded)
        if not pull or not pull.get("periods"):
            continue
        ticker = cik_to_ticker.get(cik_padded, "")
        label = ticker if ticker else cik_padded
        sheet = _safe_sheet_name(label, used_sheet_names)
        _write_peer_sheet(wb, sheet, pull, args.num_periods)
    pulled = sum(1 for v in peer_pulls.values() if v.get("periods"))
    _write_about(
        wb, args, len(peers), pulled, metric_slugs, as_of,
        args.capiq_layout, args.extensions, args.include_beta,
    )

    wb.save(out_path)
    print(f"Done. {len(peers)} peers in universe; {pulled} pulled "
          f"successfully.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
